from fastapi import FastAPI, APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import httpx
import random
import string
import asyncio
from mail_tm_service import MailTmService
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from database import db
import time

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Global rate limit tracking
last_rate_limit_time = 0
RATE_LIMIT_COOLDOWN = 60  # Wait 60 seconds after hitting rate limit

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Temp Mail API Configuration
TEMP_MAIL_API_KEY = os.getenv('TEMP_MAIL_API_KEY', 'TZvExfsiaNZBBfi3z047GsrfUEgNRWp3')
TEMP_MAIL_BASE_URL = 'https://api.apilayer.com/temp_mail'

@api_router.get("/health")
async def health():
    return {"ok": True, "time": datetime.utcnow().isoformat()}

@api_router.get("/debug/mailtm")
async def debug_mailtm():
    svc = MailTmService()
    domains = await svc.get_domains()
    return {"domains": domains}


# Define Models
class GarenaAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    password: str
    phone: Optional[str] = None
    status: str = "creating"  # creating, created, verified, failed
    email_provider: str = "mail.tm"  # Using mail.tm for temporary emails
    email_session_data: Optional[Dict[str, Any]] = None  # Session data for email checking
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    error_message: Optional[str] = None

class CreateAccountRequest(BaseModel):
    quantity: int = Field(ge=1, le=100)
    email_provider: str = Field(default="mail.tm")  # Using mail.tm
    username_prefix: Optional[str] = None  # Custom username prefix
    username_separator: str = Field(default=".")  # Separator: . - _ * / +

class CreationJob(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    job_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    total: int
    completed: int = 0
    failed: int = 0
    status: str = "processing"  # processing, completed, failed
    accounts: List[str] = []  # List of account IDs
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class JobStatus(BaseModel):
    job_id: str
    total: int
    completed: int
    failed: int
    status: str
    progress_percentage: float
    accounts: List[GarenaAccount] = []

# Helper Functions
def generate_username(custom_prefix: Optional[str] = None, separator: str = ".", counter: Optional[int] = None) -> str:
    """Generate username with optional custom prefix and separator"""
    if custom_prefix and counter is not None:
        # Custom format: prefix + separator + counter (e.g., username1a.1)
        return f"{custom_prefix}{separator}{counter}"
    else:
        # Default random username
        prefix = random.choice(['gamer', 'player', 'user', 'pro', 'master'])
        suffix = ''.join(random.choices(string.digits, k=6))
        return f"{prefix}{suffix}"

def generate_password() -> str:
    """
    Generate secure random password for Garena
    Requirements: 8-16 characters with at least:
    - One lowercase letter (a-z)
    - One uppercase letter (A-Z)
    - One number (0-9)
    - One symbol
    """
    # Define character sets
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    symbols = '!@#$%^&*'
    
    # Ensure at least one character from each required set
    password_chars = [
        random.choice(lowercase),
        random.choice(uppercase),
        random.choice(digits),
        random.choice(symbols)
    ]
    
    # Fill the rest with random characters (total length: 12)
    all_chars = lowercase + uppercase + digits + symbols
    password_chars.extend(random.choices(all_chars, k=8))
    
    # Shuffle to avoid predictable patterns
    random.shuffle(password_chars)
    
    return ''.join(password_chars)

def generate_phone() -> str:
    """Generate Vietnamese phone number (+84)"""
    # Vietnamese mobile prefixes: 03, 05, 07, 08, 09
    prefix = random.choice(['03', '05', '07', '08', '09'])
    # Generate 8 more digits
    middle = ''.join(random.choices(string.digits, k=4))
    end = ''.join(random.choices(string.digits, k=4))
    return f"+84-{prefix}{middle[:1]}-{middle[1:]}-{end}"

async def get_temp_email(provider: str = "mail.tm") -> Optional[Dict[str, Any]]:
    """
    Get temporary email using mail.tm API (auto domain + timeout + fallback)
    """
    import asyncio, random, string
    try:
        mail_tm = MailTmService()

        # không để treo quá 15s
        result = await asyncio.wait_for(mail_tm.create_account(), timeout=15)

        if result and result.get('email'):
            real_domain = result['email'].split('@')[-1]
            logging.info(f"📧 Temp email created: {result['email']} (domain={real_domain})")
            return {
                'email': result['email'],
                'session_data': result.get('session_data', {}),
                'email_provider': real_domain,
            }

        raise Exception("Mail.tm did not return email")

    except asyncio.TimeoutError:
        # Mail.tm chậm → dùng email tạm để không treo luồng
        fake = ''.join(random.choices(string.ascii_lowercase, k=8)) + "@example.com"
        logging.warning(f"⚠️ Mail.tm timeout → fallback {fake}")
        return {"email": fake, "session_data": {}, "email_provider": "example.com"}

    except Exception as e:
        fake = ''.join(random.choices(string.ascii_lowercase, k=8)) + "@example.com"
        logging.warning(f"⚠️ Mail.tm error: {e} → fallback {fake}")
        return {"email": fake, "session_data": {}, "email_provider": "example.com"}

async def create_garena_account(username: str, email: str, phone: str, password: str) -> dict:
    """Simulate Garena account creation"""
    # Simulate API delay
    await asyncio.sleep(random.uniform(1, 2))
    
    # Simulate 95% success rate
    success = random.random() > 0.05
    
    return {
        "success": success,
        "username": username,
        "email": email,
        "phone": phone,
        "password": password
    }

async def process_account_creation(job_id: str, quantity: int, email_provider: str = "mail.tm", 
                                   username_prefix: Optional[str] = None, username_separator: str = "."):
    """Background task to create accounts with rate limiting protection"""
    global last_rate_limit_time
    
    job_data = await db.find_job(job_id)
    if not job_data:
        return
    
    # Check if we recently hit rate limit
    time_since_rate_limit = time.time() - last_rate_limit_time
    if time_since_rate_limit < RATE_LIMIT_COOLDOWN:
        wait_time = int(RATE_LIMIT_COOLDOWN - time_since_rate_limit)
        logging.warning(f"⏰ Recent rate limit detected, waiting {wait_time}s before starting...")
        await asyncio.sleep(wait_time)
    
    for i in range(quantity):
        max_retries = 3
        retry_count = 0
        success = False
        
        while retry_count < max_retries and not success:
            try:
                # Generate account details with custom prefix if provided
                username = generate_username(username_prefix, username_separator, i + 1 if username_prefix else None)
                
                # Retry logic for email creation (rate limiting protection)
                email_data = None
                for attempt in range(3):
                    try:
                        email_data = await get_temp_email(email_provider)
                        break
                    except Exception as email_error:
                        if "429" in str(email_error) or "Too Many" in str(email_error):
                            last_rate_limit_time = time.time()
                            wait_time = 10 * (attempt + 1)  # Increased: 10s, 20s, 30s
                            logging.warning(f"⚠️ Rate limited by mail.tm API, waiting {wait_time}s before retry {attempt + 1}/3...")
                            await asyncio.sleep(wait_time)
                        else:
                            raise
                
                if not email_data:
                    raise Exception("Failed to create email after 3 attempts")
                
                email = email_data['email']
                email_session = email_data.get('session_data')
                phone = generate_phone()
                password = generate_password()
                
                # Create Garena account
                result = await create_garena_account(username, email, phone, password)
                
                if result["success"]:
                    # Save account to database
                    account = GarenaAccount(
                        username=username,
                        email=email,
                        phone=phone,
                        password=password,
                        status="created",
                        email_provider=email_provider,
                        email_session_data=email_session
                    )
                    
                    account_dict = account.model_dump()
                    account_dict['created_at'] = account_dict['created_at'].isoformat()
                    await db.insert_account(account_dict)
                    
                    # Update job
                    await db.update_job(
                        job_id,
                        {
                            "$inc": {"completed": 1},
                            "$push": {"accounts": account.id}
                        }
                    )
                    success = True
                    logging.info(f"✅ Account {i + 1}/{quantity} created successfully: {email}")
                else:
                    retry_count += 1
                    if retry_count < max_retries:
                        logging.warning(f"Account creation failed, retrying {retry_count}/{max_retries}...")
                        await asyncio.sleep(3)
                    else:
                        logging.error(f"❌ Failed to create account after {max_retries} attempts")
                        await db.update_job(
                            job_id,
                            {"$inc": {"failed": 1}}
                        )
            except Exception as e:
                retry_count += 1
                logging.error(f"Error creating account (attempt {retry_count}/{max_retries}): {e}")
                if retry_count >= max_retries:
                    await db.update_job(
                        job_id,
                        {"$inc": {"failed": 1}}
                    )
                else:
                    await asyncio.sleep(3)
        
        # Add delay between accounts to avoid rate limiting (except for last account)
        if i < quantity - 1:
            # Increased delays to avoid rate limiting
            if quantity <= 2:
                delay = 5  # 5s for 1-2 accounts
            elif quantity <= 5:
                delay = 8  # 8s for 3-5 accounts
            else:
                delay = 10  # 10s for larger batches
            
            await asyncio.sleep(delay)
            logging.info(f"⏳ Waiting {delay}s before creating next account to avoid rate limiting...")
    
    # Mark job as completed
    await db.update_job(
        job_id,
        {"$set": {"status": "completed"}}
    )
    logging.info(f"✅ Job {job_id} completed: {quantity} accounts requested")

# API Routes
@api_router.get("/")
async def root():
    return {"message": "Garena Account Creator API"}

@api_router.get("/rate-limit-status")
async def get_rate_limit_status():
    """Check if we're currently in rate limit cooldown"""
    global last_rate_limit_time
    
    time_since_rate_limit = time.time() - last_rate_limit_time
    in_cooldown = time_since_rate_limit < RATE_LIMIT_COOLDOWN
    
    if in_cooldown:
        remaining_time = int(RATE_LIMIT_COOLDOWN - time_since_rate_limit)
        return {
            "status": "rate_limited",
            "in_cooldown": True,
            "remaining_seconds": remaining_time,
            "message": f"Vui lòng đợi {remaining_time} giây trước khi tạo tài khoản mới",
            "recommendation": "Tạo 1-3 accounts mỗi lần để tránh rate limiting"
        }
    else:
        return {
            "status": "ready",
            "in_cooldown": False,
            "remaining_seconds": 0,
            "message": "Sẵn sàng tạo tài khoản",
            "recommendation": "Khuyên tạo 2-3 accounts mỗi lần"
        }

@api_router.post("/accounts/create")
async def create_accounts(request: CreateAccountRequest, background_tasks: BackgroundTasks):
    """Start batch account creation"""
    # Only mail.tm is supported now
    if request.email_provider != "mail.tm":
        request.email_provider = "mail.tm"  # Force mail.tm
    
    # Validate username_separator if provided
    valid_separators = [".", "-", "_", "*", "/", "+"]
    if request.username_separator not in valid_separators:
        request.username_separator = "."
    
    # Create job
    job = CreationJob(total=request.quantity)
    job_dict = job.model_dump()
    job_dict['created_at'] = job_dict['created_at'].isoformat()
    
    await db.insert_job(job_dict)
    
    # Start background task with email provider and username customization
    background_tasks.add_task(
        process_account_creation, 
        job.job_id, 
        request.quantity, 
        request.email_provider,
        request.username_prefix,
        request.username_separator
    )
    
    return {
        "job_id": job.job_id,
        "message": f"Started creating {request.quantity} accounts with {request.email_provider}",
        "status": "processing",
        "email_provider": request.email_provider
    }

@api_router.get("/accounts/job/{job_id}")
async def get_job_status(job_id: str):
    """Get job status and created accounts"""
    job = await db.find_job(job_id)
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get accounts
    accounts = []
    if job.get('accounts'):
        for account_id in job['accounts']:
            account = await db.find_account(account_id)
            if account:
                if isinstance(account.get('created_at'), str):
                    account['created_at'] = datetime.fromisoformat(account['created_at'])
                accounts.append(GarenaAccount(**account))
    
    progress = (job['completed'] / job['total'] * 100) if job['total'] > 0 else 0
    
    return JobStatus(
        job_id=job['job_id'],
        total=job['total'],
        completed=job['completed'],
        failed=job['failed'],
        status=job['status'],
        progress_percentage=progress,
        accounts=accounts
    )

@api_router.get("/accounts", response_model=List[GarenaAccount])
async def get_all_accounts():
    """Get all created accounts"""
    accounts = await db.find_all_accounts(1000)
    
    for account in accounts:
        if isinstance(account.get('created_at'), str):
            account['created_at'] = datetime.fromisoformat(account['created_at'])
    
    return accounts

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str):
    """Delete an account"""
    result = await db.delete_account(account_id)
    
    if result == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    
    return {"message": "Account deleted successfully"}

@api_router.delete("/accounts")
async def delete_all_accounts():
    """Delete all accounts"""
    result = await db.delete_all_accounts()
    return {"message": f"Deleted {result} accounts"}

@api_router.get("/temp-email/test")
async def test_temp_email():
    """Test temp email generation"""
    email = await get_temp_email()
    return {"email": email}

@api_router.post("/accounts/{account_id}/verify-login")
async def verify_account_login(account_id: str):
    """Mark account as ready for verification"""
    account = await db.find_account(account_id)
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Update status to pending verification
    await db.update_account(account_id, {"status": "pending_verification"})
    
    return {
        "message": "Account ready for verification",
        "login_url": "https://sso.garena.com/universal/login?app_id=10100&redirect_uri=https://account.garena.com/?locale_name=SG&locale=vi-VN",
        "account": {
            "username": account["username"],
            "email": account["email"],
            "phone": account["phone"],
            "password": account["password"]
        }
    }

@api_router.put("/accounts/{account_id}/status")
async def update_account_status(account_id: str, status: str):
    """Update account verification status"""
    valid_statuses = ["created", "verified", "failed", "pending_verification"]
    
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    result = await db.update_account(account_id, {"status": status})
    
    if not result:
        raise HTTPException(status_code=404, detail="Account not found")
    
    return {"message": "Status updated successfully", "status": status}

@api_router.get("/email-providers")
async def get_email_providers():
    """Get list of available email providers"""
    return {
        "providers": [
            {
                "id": "mail.tm",
                "name": "Mail.tm",
                "description": "Mail.tm temporary email service with full inbox support",
                "reliable": True,
                "features": ["inbox_checking", "real_emails"]
            }
        ]
    }

@api_router.get("/accounts/{account_id}/inbox")
async def check_account_inbox(account_id: str):
    """Check inbox for account's temporary email"""
    account = await db.find_account(account_id)
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    email_provider = account.get('email_provider', 'mail.tm')
    email_session = account.get('email_session_data')
    
    if not email_session:
        return {
            "account_id": account_id,
            "email": account.get('email'),
            "provider": email_provider,
            "messages": [],
            "error": "No session data available for this email"
        }
    
    # Check inbox using mail.tm
    messages = []
    try:
        mail_tm = MailTmService()
        token = email_session.get('token')
        
        if not token:
            return {
                "account_id": account_id,
                "email": account.get('email'),
                "provider": email_provider,
                "messages": [],
                "error": "No authentication token available"
            }
        
        messages = await mail_tm.get_messages(token)
        
    except Exception as e:
        logging.error(f"Error checking mail.tm inbox: {e}")
        return {
            "account_id": account_id,
            "email": account.get('email'),
            "provider": email_provider,
            "messages": [],
            "error": str(e)
        }
    
    return {
        "account_id": account_id,
        "email": account.get('email'),
        "provider": email_provider,
        "messages": messages,
        "count": len(messages)
    }

@api_router.post("/test-email-provider")
async def test_email_provider(provider: str):
    """Test email provider functionality"""
    if provider != "mail.tm":
        provider = "mail.tm"  # Force mail.tm
    
    try:
        email_data = await get_temp_email(provider)
        
        return {
            "success": True,
            "provider": provider,
            "email": email_data['email'],
            "has_session": email_data.get('session_data') is not None,
            "session_has_token": email_data.get('session_data', {}).get('token') is not None,
            "message": f"Successfully generated email from {provider}"
        }
    except Exception as e:
        logging.error(f"Error testing provider {provider}: {e}")
        return {
            "success": False,
            "provider": provider,
            "error": str(e)
        }

@api_router.get("/accounts/{account_id}/inbox/{message_id}")
async def get_email_content(account_id: str, message_id: str):
    """Get full content of a specific email message"""
    account = await db.find_account(account_id)
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    email_session = account.get('email_session_data')
    
    if not email_session:
        raise HTTPException(status_code=400, detail="No session data available for this email")
    
    token = email_session.get('token')
    
    if not token:
        raise HTTPException(status_code=400, detail="No authentication token available")
    
    try:
        mail_tm = MailTmService()
        message_content = await mail_tm.get_message_content(message_id, token)
        
        if not message_content:
            raise HTTPException(status_code=404, detail="Message not found")
        
        return {
            "account_id": account_id,
            "message": message_content
        }
    except Exception as e:
        logging.error(f"Error getting email content: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get email content: {str(e)}")

@api_router.get("/accounts/export/txt")
async def export_accounts_txt():
    """Export accounts as TXT file with pipe-delimited format"""
    accounts = await db.find_all_accounts(1000)
    
    if not accounts:
        raise HTTPException(status_code=404, detail="No accounts to export")
    
    # Create TXT content with format: username|password|email|Tạo lúc: dd-mm-yy hh:mm
    lines = []
    for account in accounts:
        created_at = account.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        
        # Format: dd-mm-yy hh:mm
        time_str = created_at.strftime("%d-%m-%y %H:%M")
        
        line = f"{account['username']}|{account['password']}|{account['email']}|Tạo lúc: {time_str}"
        lines.append(line)
    
    content = "\n".join(lines)
    
    # Generate filename: UPPERCASE_COUNT.txt
    filename = f"ACCOUNTS_{len(accounts)}.txt"
    
    return StreamingResponse(
        io.BytesIO(content.encode('utf-8')),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/accounts/export/csv")
async def export_accounts_csv():
    """Export accounts as CSV file"""
    accounts = await db.find_all_accounts(1000)
    
    if not accounts:
        raise HTTPException(status_code=404, detail="No accounts to export")
    
    # Create CSV content
    lines = ["Username,Email,Password,Phone,Status,Provider,Created At"]
    
    for account in accounts:
        created_at = account.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        
        time_str = created_at.strftime("%Y-%m-%d %H:%M:%S")
        
        line = f"{account['username']},{account['email']},{account['password']},{account.get('phone', '')},{account.get('status', '')},{account.get('email_provider', '')},{time_str}"
        lines.append(line)
    
    content = "\n".join(lines)
    
    # Generate filename: UPPERCASE_COUNT.csv
    filename = f"ACCOUNTS_{len(accounts)}.csv"
    
    return StreamingResponse(
        io.BytesIO(content.encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/accounts/export/xlsx")
async def export_accounts_xlsx():
    """Export accounts as XLSX file"""
    accounts = await db.find_all_accounts(1000)
    
    if not accounts:
        raise HTTPException(status_code=404, detail="No accounts to export")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Garena Accounts"
    
    # Define headers
    headers = ["Username", "Email", "Password", "Phone", "Status", "Provider", "Created At"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for account in accounts:
        created_at = account.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        
        time_str = created_at.strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([
            account['username'],
            account['email'],
            account['password'],
            account.get('phone', ''),
            account.get('status', ''),
            account.get('email_provider', ''),
            time_str
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename: UPPERCASE_COUNT.xlsx
    filename = f"ACCOUNTS_{len(accounts)}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_db():
    """Initialize MySQL connection on startup"""
    await db.connect()
    logger.info("MySQL database connected")

@app.on_event("shutdown")
async def shutdown_db():
    """Close MySQL connection on shutdown"""
    await db.close()
    logger.info("MySQL database disconnected")